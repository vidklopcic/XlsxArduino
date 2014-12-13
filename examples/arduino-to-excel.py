import time
import serial
import openpyxl
import openpyxl.cell


class ArduinoToXlsx():
    def __init__(self, port, worksheet_title, xlsx_filename, 
                 delimiter='|', left_empty_cols=1, top_empty_lines=1, 
                 time_format='%Y.%m.%d - %H:%M:%S', baud_rate=9600):
        self.left_empty_cols = left_empty_cols
        self.top_empty_lines = top_empty_lines
        self.worksheet_title = worksheet_title
        self.delimiter = delimiter
        self.time_format = time_format
        self.serial = serial.Serial(port, baud_rate)
        self.xlsx_filename = xlsx_filename

        try:
            self.xlsx_file = openpyxl.load_workbook(self.xlsx_filename)
            try:
                self.xlsx = self.xlsx_file[self.worksheet_title]
            except:
                self.xlsx_file.create_sheet(title=self.worksheet_title)
                self.xlsx = self.xlsx_file[self.worksheet_title]
        except:
            self.xlsx_file = openpyxl.Workbook()
            self.xlsx = self.xlsx_file.active
            self.xlsx.title = self.worksheet_title
        self.save()
        self.column_ending = 1
        column_letter = self.get_col_letter(0)
        for x in range(1, 10000, 1):
            cell_value = self.xlsx['%s%s' % (column_letter, x+self.top_empty_lines)].value
            if cell_value is None:
                self.column_ending = x
                break
        self.xlsx.column_dimensions[self.get_col_letter(0)].width = 20

    def recieve_serial(self):
        msg = self.serial.readline().strip().split(self.delimiter)
        self.write_msg(msg)

    def get_col_letter(self, i):
        return openpyxl.cell.get_column_letter(i+self.left_empty_cols+1)

    def write_msg(self, msg):
        for i in range(len(msg)):
            column_letter = self.get_col_letter(i+1)
            self.xlsx['%s%s' % (column_letter, self.column_ending+self.top_empty_lines)].value = msg[i]
        print msg
        self.xlsx['%s%s' % (self.get_col_letter(0), self.column_ending+self.top_empty_lines)].value = time.strftime(self.time_format)
        self.column_ending += 1

    def save(self):
        self.xlsx_file.save(self.xlsx_filename)

if __name__ == '__main__':
    atx = ArduinoToXlsx('COM3', 'worksheet1', 'test.xlsx')
    while 1:
        atx.recieve_serial()
        atx.save()
